from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook

from .models import Region, City, Building


# =======================================================
#   🟩 قائمة المناطق
# =======================================================
@login_required
def regions_list_view(request):
    regions = Region.objects.all().order_by("name")
    return render(request, "locations_app/regions_list.html", {"regions": regions})


# =======================================================
#   🟦 قائمة المدن
# =======================================================
@login_required
def cities_list_view(request):
    cities = City.objects.select_related("region").all().order_by("region__name", "name")
    return render(request, "locations_app/cities_list.html", {"cities": cities})


# =======================================================
#   🟧 قائمة المباني
# =======================================================
@login_required
def buildings_list_view(request):
    buildings = Building.objects.select_related("city", "city__region").all().order_by(
        "city__region__name", "city__name", "name"
    )
    return render(request, "locations_app/buildings_list.html", {"buildings": buildings})


# =======================================================
#   🟢 استيراد المناطق (Regions)
#       Excel structure:
#       | region_name |
# =======================================================
@login_required
def admin_import_regions(request):

    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "الرجاء اختيار ملف Excel.")
            return redirect("locations_app:admin_import_regions")

        try:
            wb = load_workbook(file)
            ws = wb.active

            added = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                region_name = row[0]

                if not region_name:
                    continue

                _, created = Region.objects.get_or_create(name=str(region_name).strip())
                if created:
                    added += 1

            messages.success(request, f"تم استيراد {added} منطقة بنجاح!")

        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء قراءة الملف: {e}")

        return redirect("inventory_app:admin_dashboard")

    return render(request, "locations_app/import_regions.html")


# =======================================================
#   🟦 استيراد المدن (Cities)
#       Excel structure:
#       | region_name | city_name |
# =======================================================
@login_required
def admin_import_cities(request):

    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "الرجاء اختيار ملف Excel.")
            return redirect("locations_app:admin_import_cities")

        try:
            wb = load_workbook(file)
            ws = wb.active

            added = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                region_name = row[0]
                city_name = row[1]

                if not region_name or not city_name:
                    continue

                region, _ = Region.objects.get_or_create(name=str(region_name).strip())
                _, created = City.objects.get_or_create(
                    region=region,
                    name=str(city_name).strip()
                )

                if created:
                    added += 1

            messages.success(request, f"تم استيراد {added} مدينة بنجاح!")

        except Exception as e:
            messages.error(request, f"فشل الاستيراد: {e}")

        return redirect("inventory_app:admin_dashboard")

    return render(request, "locations_app/import_cities.html")


# =======================================================
#   🟧 استيراد المباني (Buildings)
#       Excel structure:
#       | city_name | building_name |
# =======================================================
@login_required
def admin_import_buildings(request):

    if request.method == "POST":
        file = request.FILES.get("file")

        if not file:
            messages.error(request, "الرجاء اختيار ملف Excel.")
            return redirect("locations_app:admin_import_buildings")

        try:
            wb = load_workbook(file)
            ws = wb.active

            added = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                city_name = row[0]
                building_name = row[1]

                if not city_name or not building_name:
                    continue

                city, _ = City.objects.get_or_create(name=str(city_name).strip())
                _, created = Building.objects.get_or_create(
                    city=city,
                    name=str(building_name).strip()
                )

                if created:
                    added += 1

            messages.success(request, f"تم استيراد {added} مبنى بنجاح!")

        except Exception as e:
            messages.error(request, f"فشل الاستيراد: {e}")

        return redirect("inventory_app:admin_dashboard")

    return render(request, "locations_app/import_buildings.html")

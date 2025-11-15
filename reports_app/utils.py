import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


def generate_excel(headers, rows, filename="report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active

    # كتابة العناوين
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # كتابة الصفوف
    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center")

    # تجهيز الملف للتحميل
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

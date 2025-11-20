from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.template.loader import render_to_string
from django.db import models, transaction
from django.contrib import messages

import openpyxl
from openpyxl.styles import Font, Alignment

from locations_app.models import Region, City, Building
from assets_app.models import Asset
from inventory_app.models import InventorySession, InventoryItem



# ============================================================
# الصلاحيات
# ============================================================
def is_employee(user):
    return user.groups.filter(name="employees").exists()

def is_supervisor(user):
    return user.groups.filter(name="supervisors").exists()

def is_admin(user):
    return user.is_superuser or user.groups.filter(name="admins").exists()



# ============================================================
# الموظف — قائمة الجلسات
# ============================================================
@login_required
def sessions_list_view(request):
    sessions = InventorySession.objects.filter(
        employee=request.user
    ).order_by("-start_time")

    return render(request, "inventory_app/sessions_list.html", {
        "sessions": sessions,
    })



# ============================================================
# تفاصيل الجلسة
# ============================================================
@login_required
def session_detail_view(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)
    items = InventoryItem.objects.filter(session=session).select_related("asset")

    if (
        session.employee != request.user
        and not is_supervisor(request.user)
        and not is_admin(request.user)
    ):
        return HttpResponseForbidden("غير مصرح لك")

    return render(request, "inventory_app/session_detail.html", {
        "session": session,
        "items": items,
    })



# ============================================================
# 🔵 API — جلب المدن حسب المنطقة
# ============================================================
@login_required
def get_cities_by_region(request, region_id):
    cities = City.objects.filter(region_id=region_id).values("id", "name")
    return JsonResponse(list(cities), safe=False)



# ============================================================
# 🔵 API — جلب المباني حسب المدينة
# ============================================================
@login_required
def get_buildings_by_city(request, city_id):
    buildings = Building.objects.filter(city_id=city_id).values("id", "name")
    return JsonResponse(list(buildings), safe=False)



# ============================================================
# بدء جلسة جرد (موظف أو مدير)
# ============================================================
@login_required
def start_session_view(request):
    if not is_employee(request.user) and not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    regions = Region.objects.all()
    cities = City.objects.all()
    buildings = Building.objects.all()

    if request.method == "POST":
        region = get_object_or_404(Region, id=request.POST.get("region"))
        city = get_object_or_404(City, id=request.POST.get("city"))
        building = get_object_or_404(Building, id=request.POST.get("building"))

        session = InventorySession.objects.create(
            employee=request.user,
            region=region,
            city=city,
            building=building,
            status="in_progress",
        )

        for asset in Asset.objects.filter(region=region, city=city, building=building):
            InventoryItem.objects.create(
                session=session,
                asset=asset,
                barcode=asset.barcode,
                status="missing",
            )

        return redirect("inventory_app:live_scan", session_id=session.id)

    return render(request, "inventory_app/start_session.html", {
        "regions": regions,
        "cities": cities,
        "buildings": buildings,
    })



# ============================================================
# شاشة المسح
# ============================================================
@login_required
def live_scan_view(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    items = InventoryItem.objects.filter(session=session).select_related("asset")
    total_items = items.count()
    count_found = items.filter(status="found").count()
    count_remaining = total_items - count_found

    return render(request, "inventory_app/session_live_scan.html", {
        "session": session,
        "items": items,
         "total_items": total_items,
        "show_copy_button": False,  # نسخ الأصل فقط في شاشة الإضافة
        "count_found": count_found,
        "count_remaining": count_remaining,
})




# ============================================================
# API — تسجيل المسح (مُحسّن)
# ============================================================
@login_required
@require_POST
def scan_update_api(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    barcode = request.POST.get("barcode", "").strip()

    # 1) هل الباركود موجود في جلسة الجرد؟
    item = InventoryItem.objects.filter(session=session, barcode=barcode).first()

    if item:
        item.status = "found"
        item.scanned_at = timezone.now()
        item.save()

        return JsonResponse({
            "status": "found",
            "barcode": barcode,
            "description": item.asset.description if item.asset else "",
        })

    # 2) هل الأصل موجود في النظام لكن لم يتم إضافته للجلسة؟
    asset = Asset.objects.filter(barcode=barcode).first()

    if asset:
        # لم يكن في الجلسة — نضيفه تلقائياً
        new_item = InventoryItem.objects.create(
            session=session,
            asset=asset,
            barcode=barcode,
            status="found",
            scanned_at=timezone.now(),
            added_manually=False
        )

        return JsonResponse({
            "status": "found_new_in_system",
            "barcode": barcode,
            "description": asset.description,
        })

    # 3) باركود غير موجود نهائيًا → يعرض شاشة إضافة أصل جديد
    return JsonResponse({
        "status": "not_in_list",
        "barcode": barcode
    })

@login_required
@require_POST
def manual_confirm_api(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    # قراءة البيانات القادمة من fetch
    import json
    data = json.loads(request.body.decode("utf-8"))
    barcode = data.get("barcode")

    item = InventoryItem.objects.filter(session=session, barcode=barcode).first()

    if not item:
        return JsonResponse({"status": "not_found"}, status=404)

    # تأكيد أصل موجود يدويًا
    item.status = "found"
    item.scanned_at = timezone.now()
    item.save()

    return JsonResponse({"status": "ok"})



# ============================================================
# API — جلب بيانات أصل حسب الباركود (لاستخدام النسخ)
# ============================================================
@login_required
def get_asset_api(request, barcode):
    """
    API – جلب بيانات أصل كامل
    تستخدم في:
    1) نافذة معاينة الأصل عند الضغط على الباركود
    2) نظام نسخ أصل مشابه (Copy Asset)
    """
    try:
        asset = Asset.objects.select_related("region", "city", "building").get(barcode=barcode)
    except Asset.DoesNotExist:
        return JsonResponse({
            "status": "not_found",
            "message": "الأصل غير موجود في قاعدة البيانات"
        })

    return JsonResponse({
        "status": "found",

        # بيانات عامة
        "asset_code": asset.asset_code,
        "barcode": asset.barcode,
        "old_barcode": asset.old_barcode,

        # الوصف والتصنيفات
        "description": asset.description,
        "main_category": asset.main_category,
        "type": asset.type,
        "sub_category": asset.sub_category,

        # الموقع
        "region": asset.region.name if asset.region else "",
        "city": asset.city.name if asset.city else "",
        "building": asset.building.name if asset.building else "",

        # الحالة والعهدة
        "status_text": asset.status,
        "condition": asset.condition,
        "custodian_name": asset.custodian_name,
        "custodian_number": asset.custodian_number,
        "custodian_type": asset.custodian_type,
        "phone_number": asset.phone_number,

        # إنشاء
        "created_at": str(asset.created_at),
        "created_by": asset.created_by_username,

    })


# ============================================================
# API — إضافة أصل جديد
# ============================================================
@login_required
@require_POST
def add_new_asset_api(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    # لو جاي من النسخ
    source_barcode = request.POST.get("source_barcode")
    new_barcode = request.POST.get("new_barcode")

    if source_barcode and new_barcode:
        source = Asset.objects.filter(barcode=source_barcode).first()
        if not source:
            return JsonResponse({"status": "not_found"})

        new_asset = Asset.objects.create(
            asset_code=source.asset_code,
            barcode=new_barcode,
            old_barcode=source.old_barcode,
            description=source.description,
            main_category=source.main_category,
            type=source.type,
            sub_category=source.sub_category,
            region=session.region,
            city=session.city,
            building=session.building,
            status=source.status,
            condition=source.condition,
            custodian_name=request.POST.get("custodian_name"),
            custodian_number=request.POST.get("custodian_number"),
            custodian_type=request.POST.get("custodian_type"),
            created_at=timezone.now(),
            created_by_username=request.user.username,
        )

        InventoryItem.objects.create(
            session=session,
            asset=new_asset,
            barcode=new_barcode,
            status="new",
            added_manually=True
        )

        return JsonResponse({"status": "success"})

    # إضافة أصل جديد بدون نسخ
    barcode = request.POST.get("barcode")
    description = request.POST.get("description")

    asset = Asset.objects.create(
        asset_code=request.POST.get("asset_code"),
        barcode=barcode,
        description=description,
        main_category=request.POST.get("category"),
        sub_category=request.POST.get("subcategory"),
        condition=request.POST.get("condition"),
        region=session.region,
        city=session.city,
        building=session.building,
        type="غير محدد",
        created_at=timezone.now(),
        created_by_username=request.user.username,
    )

    InventoryItem.objects.create(
        session=session,
        asset=asset,
        barcode=barcode,
        status="new",
        added_manually=True
    )

    return JsonResponse({"status": "new_added"})



# ============================================================
# إنهاء الجلسة
# ============================================================
@login_required
@require_POST
def close_session(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    # تحويل الجلسة بدلاً من إغلاقها
    session.status = "submitted_to_supervisor"
    session.end_time = timezone.now()
    session.save()

    return JsonResponse({"status": "success"})




# ============================================================
# إرسال للمشرف
# ============================================================
@login_required
@require_POST
def submit_to_supervisor(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    session.status = "submitted_to_supervisor"
    session.save()

    return JsonResponse({"status": "success"})



# ============================================================
# المشرف — قائمة الجلسات
# ============================================================
@login_required
def supervisor_sessions_list(request):
    if not is_supervisor(request.user) and not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    sessions = InventorySession.objects.filter(status="submitted_to_supervisor")

    return render(request, "inventory_app/supervisor_sessions_list.html", {
        "sessions": sessions,
    })



# ============================================================
# المشرف — تفاصيل الجلسة
# ============================================================
@login_required
def supervisor_session_detail(request, session_id):
    if not is_supervisor(request.user) and not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    session = get_object_or_404(InventorySession, id=session_id)
    items = InventoryItem.objects.filter(session=session)

    return render(request, "inventory_app/supervisor_session_detail.html", {
        "session": session,
        "items": items,
    })



# ============================================================
# المشرف — موافقة
# ============================================================
@login_required
@require_POST
def supervisor_approve_session(request, session_id):
    if not is_supervisor(request.user) and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    session = get_object_or_404(InventorySession, id=session_id)

    session.status = "supervisor_approved"
    session.supervisor = request.user
    session.save()

    return JsonResponse({"status": "success"})



# ============================================================
# المشرف — رفض
# ============================================================
@login_required
@require_POST
def supervisor_reject_session(request, session_id):
    if not is_supervisor(request.user) and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    session = get_object_or_404(InventorySession, id=session_id)
    comment = request.POST.get("comment", "").strip()

    session.status = "supervisor_rejected"
    session.supervisor = request.user
    session.supervisor_comment = comment
    session.save()

    return JsonResponse({"status": "success"})



# ============================================================
# المدير — قائمة الجلسات
# ============================================================
@login_required
def admin_sessions_list(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    sessions = InventorySession.objects.all().order_by("-start_time")

    return render(request, "inventory_app/admin_sessions_list.html", {
        "sessions": sessions,
    })



# ============================================================
# المدير — تفاصيل الجلسة
# ============================================================
@login_required
def admin_session_detail(request, session_id):

    if not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    session = get_object_or_404(InventorySession, id=session_id)
    items = InventoryItem.objects.filter(session=session)

    return render(request, "inventory_app/admin_session_detail.html", {
        "session": session,
        "items": items,
    })



# ============================================================
# المدير — إعادة فتح الجلسة
# ============================================================
@login_required
@require_POST
def admin_reopen_session(request, session_id):

    if not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    session = get_object_or_404(InventorySession, id=session_id)

    session.status = "in_progress"
    session.end_time = None
    session.supervisor = None
    session.supervisor_comment = ""
    session.save()

    return JsonResponse({"status": "success"})



# ============================================================
# المدير — حذف الجلسة
# ============================================================
@login_required
@require_POST
def admin_delete_session(request, session_id):

    if not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    try:
        session = InventorySession.objects.get(id=session_id)
        session.delete()
        return JsonResponse({"status": "success"})

    except InventorySession.DoesNotExist:
        return JsonResponse({"status": "error", "message": "الجلسة غير موجودة"})



# ============================================================
# تصدير PDF
# ============================================================
@login_required
def export_session_pdf(request, session_id):

    session = get_object_or_404(InventorySession, id=session_id)
    items = InventoryItem.objects.filter(session=session).select_related("asset")

    html = render_to_string("inventory_app/report_template.html", {
        "session": session,
        "items": items,
    })

    from xhtml2pdf import pisa
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = f'attachment; filename="session_{session_id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("خطأ أثناء إنشاء ملف PDF", status=500)

    return response



# ============================================================
# تصدير Excel
# ============================================================
@login_required
def export_session_excel(request, session_id):

    session = get_object_or_404(InventorySession, id=session_id)
    items = InventoryItem.objects.filter(session=session).select_related("asset")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Session"

    headers = ["الباركود", "الوصف", "الحالة", "وقت المسح"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([
            item.barcode,
            item.asset.description,
            item.status,
            item.scanned_at.strftime("%Y-%m-%d %H:%M") if item.scanned_at else "-",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="session_{session_id}.xlsx"'

    wb.save(response)
    return response



# ===========================
# مدير النظام — لوحة التحكم Dashboard
# ===========================
@login_required
def admin_dashboard(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك (Admin فقط)")

    total_sessions = InventorySession.objects.count()
    completed = InventorySession.objects.filter(status="completed").count()
    submitted = InventorySession.objects.filter(status="submitted_to_supervisor").count()
    approved = InventorySession.objects.filter(status="supervisor_approved").count()
    rejected = InventorySession.objects.filter(status="supervisor_rejected").count()

    latest_sessions = InventorySession.objects.select_related(
        "employee", "region", "building"
    ).order_by("-start_time")[:5]

    top_users = InventorySession.objects.values("employee__username") \
        .annotate(count=models.Count("id")) \
        .order_by("-count")[:5]

    return render(request, "inventory_app/admin_dashboard.html", {
        "total_sessions": total_sessions,
        "completed": completed,
        "submitted": submitted,
        "approved": approved,
        "rejected": rejected,
        "latest_sessions": latest_sessions,
        "top_users": top_users,
    })



# ===========================
# مدير النظام — حذف الجلسة
# ===========================
@login_required
@require_POST
def admin_delete_session(request, session_id):

    if not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    try:
        session = InventorySession.objects.get(id=session_id)
        session.delete()
        return JsonResponse({"status": "success"})

    except InventorySession.DoesNotExist:
        return JsonResponse({"status": "error", "message": "الجلسة غير موجودة"})



# ============================================================
# استيراد الأصول
# ============================================================
@login_required
def admin_import_assets(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("غير مسموح لك")

    if request.method == "POST" and request.FILES.get("excel_file"):

        file = request.FILES["excel_file"]

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except:
            messages.error(request, "❌ خطأ: الملف غير صالح، تأكد أنه Excel بصيغة .xlsx")
            return redirect("inventory_app:admin_import_assets")

        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        col = {name: index for index, name in enumerate(headers)}

        required_cols = [
            "asset_code", "barcode", "old_barcode", "description",
            "main_category", "type", "sub_category",
            "region_name", "city_name", "building_name",
            "status", "condition",
            "custodian_number", "custodian_name", "custodian_type",
            "created_at", "created_by_username",
        ]

        missing = [c for c in required_cols if c not in col]
        if missing:
            messages.error(request, f"❌ الأعمدة المفقودة: {', '.join(missing)}")
            return redirect("inventory_app:admin_import_assets")

        added = 0
        skipped = 0
        errors = []

        with transaction.atomic():
            for row_number, row in enumerate(ws.iter_rows(min_row=2), start=2):

                try:
                    get = lambda field: row[col[field]].value if col[field] < len(row) else None

                    asset_code       = get("asset_code")
                    barcode          = get("barcode")
                    old_barcode      = get("old_barcode")
                    description      = get("description")
                    main_category    = get("main_category")
                    type_            = get("type")
                    sub_category     = get("sub_category")
                    region_name      = get("region_name")
                    city_name        = get("city_name")
                    building_name    = get("building_name")
                    status           = get("status")
                    condition        = get("condition")
                    custodian_number = get("custodian_number")
                    custodian_name   = get("custodian_name")
                    custodian_type   = get("custodian_type")
                    created_at       = get("created_at")
                    created_by_username = get("created_by_username")

                    if not asset_code or not barcode:
                        skipped += 1
                        errors.append(f"سطر {row_number}: asset_code أو barcode فارغ")
                        continue

                    region = Region.objects.filter(name=str(region_name).strip()).first()
                    city = City.objects.filter(name=str(city_name).strip()).first()
                    building = Building.objects.filter(name=str(building_name).strip()).first()

                    if not region or not city or not building:
                        skipped += 1
                        errors.append(f"سطر {row_number}: بيانات الموقع غير صحيحة")
                        continue

                    Asset.objects.create(
                        asset_code=asset_code,
                        barcode=barcode,
                        old_barcode=old_barcode,
                        description=description,
                        main_category=main_category,
                        type=type_,
                        sub_category=sub_category,
                        region=region,
                        city=city,
                        building=building,
                        status=status,
                        condition=condition,
                        custodian_number=custodian_number,
                        custodian_name=custodian_name,
                        custodian_type=custodian_type,
                        created_at=created_at,
                        created_by_username=created_by_username,
                    )

                    added += 1

                except Exception as e:
                    skipped += 1
                    errors.append(f"سطر {row_number}: {str(e)}")

        return render(request, "inventory_app/admin_import_result.html", {
            "added": added,
            "skipped": skipped,
            "errors": errors,
        })

    return render(request, "inventory_app/admin_import_assets.html")



# ============================================================
# النسخة الاحتياطية الكاملة
# ============================================================
@login_required
def backup_full_system(request):
    if not is_admin(request.user):
        return HttpResponseForbidden("غير مصرح لك")

    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1 — ملخص الجلسات
    # ============================================================
    ws1 = wb.active
    ws1.title = "Sessions_Summary"

    headers1 = [
        "session_id", "employee",
        "region", "city", "building",
        "status", "start_time", "end_time",
        "total_items", "found_items",
        "missing_items", "new_items",
    ]
    ws1.append(headers1)

    sessions = InventorySession.objects.select_related(
        "employee", "region", "city", "building"
    )

    for session in sessions:
        items = InventoryItem.objects.filter(session=session)

        ws1.append([
            session.id,
            session.employee.username if session.employee else "",
            session.region.name if session.region else "",
            session.city.name if session.city else "",
            session.building.name if session.building else "",
            session.status,
            session.start_time.strftime("%Y-%m-%d %H:%M") if session.start_time else "",
            session.end_time.strftime("%Y-%m-%d %H:%M") if session.end_time else "",
            items.count(),
            items.filter(status="found").count(),
            items.filter(status="missing").count(),
            items.filter(status="new").count(),
        ])

    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


    # ============================================================
    # Sheet 2 — تفاصيل العناصر
    # ============================================================
    ws2 = wb.create_sheet(title="Items_Details")

    headers2 = [
        "session_id", "asset_code", "barcode",
        "description", "status", "scanned_at",
        "region", "city", "building",
    ]
    ws2.append(headers2)

    items = InventoryItem.objects.select_related(
        "session", "asset",
        "asset__region", "asset__city", "asset__building"
    )

    for item in items:
        ws2.append([
            item.session.id,
            item.asset.asset_code if item.asset else "",
            item.barcode,
            item.asset.description if item.asset else "",
            item.status,
            item.scanned_at.strftime("%Y-%m-%d %H:%M") if item.scanned_at else "",
            item.asset.region.name if item.asset and item.asset.region else "",
            item.asset.city.name if item.asset and item.asset.city else "",
            item.asset.building.name if item.asset and item.asset.building else "",
        ])

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")



# ============================================================
# تحميل النسخة الاحتياطية الكاملة
# ============================================================
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="full_inventory_backup.xlsx"'

    wb.save(response)
    return response



# ============================================================
#    الحفظ المؤقت للجلسة
# ============================================================
@login_required
@require_POST
def save_draft_session(request, session_id):
    session = get_object_or_404(InventorySession, id=session_id)

    if session.employee != request.user and not is_admin(request.user):
        return JsonResponse({"status": "forbidden"}, status=403)

    # فقط نحفظ بدون تغيير الحالة
    session.status = "draft"
    session.save()

    return JsonResponse({"status": "success"})




